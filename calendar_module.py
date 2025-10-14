from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from models import db, Schedule, Group, Lesson, Attendance
from ai_utils import AIAnalyzer
from datetime import datetime, timedelta
import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

calendar_bp = Blueprint('calendar', __name__)
ai = AIAnalyzer()


def parse_schedule_excel_with_mapping(file_path, column_mapping, start_row, file_extension='.xlsx'):
    """
    Парсит Excel файл с расписанием используя настройки колонок
    """
    try:
        # Читаем Excel файл
        if file_extension == '.xlsx':
            df = pd.read_excel(file_path, sheet_name=0, header=None, engine='openpyxl')
        else:
            df = pd.read_excel(file_path, sheet_name=0, header=None, engine='xlrd')
        
        lessons = []
        
        # Проверяем обязательные колонки
        required_columns = ['title', 'group', 'date', 'time']
        for col in required_columns:
            if col not in column_mapping or column_mapping[col] == '':
                raise ValueError(f'Не указана колонка для поля: {col}')
        
        # Ищем строки с данными начиная с указанной строки
        for index, row in df.iterrows():
            if index < start_row:  # Пропускаем строки до начала данных
                continue
                
            # Извлекаем данные по указанным колонкам
            try:
                title = str(row.iloc[int(column_mapping['title'])]).strip() if pd.notna(row.iloc[int(column_mapping['title'])]) else ""
                group = str(row.iloc[int(column_mapping['group'])]).strip() if pd.notna(row.iloc[int(column_mapping['group'])]) else ""
                date_str = str(row.iloc[int(column_mapping['date'])]).strip() if pd.notna(row.iloc[int(column_mapping['date'])]) else ""
                time_str = str(row.iloc[int(column_mapping['time'])]).strip() if pd.notna(row.iloc[int(column_mapping['time'])]) else ""
                
                # Аудитория (опционально)
                classroom = ""
                if 'classroom' in column_mapping and column_mapping['classroom'] != '':
                    classroom = str(row.iloc[int(column_mapping['classroom'])]).strip() if pd.notna(row.iloc[int(column_mapping['classroom'])]) else ""
                
                # Пропускаем пустые строки
                if not title or not group or not date_str or not time_str:
                    continue
                    
                # Парсим дату
                try:
                    date_obj = pd.to_datetime(date_str, format='%d.%m.%Y').date()
                except:
                    try:
                        date_obj = pd.to_datetime(date_str).date()
                    except:
                        continue
                        
                # Парсим время
                try:
                    if '-' in time_str:
                        start_time_str, end_time_str = time_str.split('-')
                        start_time = datetime.strptime(f"{date_obj} {start_time_str.strip()}", "%Y-%m-%d %H.%M")
                        end_time = datetime.strptime(f"{date_obj} {end_time_str.strip()}", "%Y-%m-%d %H.%M")
                    else:
                        continue
                except:
                    continue
                    
                # Создаем название занятия
                lesson_title = f"{title}"
                
                lessons.append({
                    'title': lesson_title,
                    'group': group,
                    'date': date_obj,
                    'start_time': start_time,
                    'end_time': end_time,
                    'classroom': classroom
                })
                
            except (IndexError, ValueError) as e:
                # Пропускаем строки с ошибками
                continue
                
        return lessons
        
    except Exception as e:
        print(f"Error parsing Excel file: {e}")
        return []


def parse_schedule_excel(file_path):
    """
    Парсит Excel файл с расписанием и извлекает данные о занятиях
    """
    try:
        # Читаем Excel файл
        df = pd.read_excel(file_path, sheet_name='Лист1', header=None)
        
        lessons = []
        
        # Ищем строки с данными о занятиях (начиная с 6-й строки, где есть данные)
        for index, row in df.iterrows():
            if index < 5:  # Пропускаем заголовки
                continue
                
            # Проверяем, есть ли данные в строке
            if pd.isna(row[0]) or pd.isna(row[2]) or pd.isna(row[7]) or pd.isna(row[9]) or pd.isna(row[10]) or pd.isna(row[12]):
                continue
                
            # Извлекаем данные
            lesson_type = str(row[0]).strip() if not pd.isna(row[0]) else ""
            discipline = str(row[2]).strip() if not pd.isna(row[2]) else ""
            date_str = str(row[7]).strip() if not pd.isna(row[7]) else ""
            time_str = str(row[9]).strip() if not pd.isna(row[9]) else ""
            classroom = str(row[10]).strip() if not pd.isna(row[10]) else ""
            group = str(row[12]).strip() if not pd.isna(row[12]) else ""
            
            # Пропускаем пустые строки
            if not discipline or not date_str or not time_str or not group:
                continue
                
            # Парсим дату
            try:
                date_obj = pd.to_datetime(date_str, format='%d.%m.%Y').date()
            except:
                try:
                    date_obj = pd.to_datetime(date_str).date()
                except:
                    continue
                    
            # Парсим время
            try:
                if '-' in time_str:
                    start_time_str, end_time_str = time_str.split('-')
                    start_time = datetime.strptime(f"{date_obj} {start_time_str.strip()}", "%Y-%m-%d %H.%M")
                    end_time = datetime.strptime(f"{date_obj} {end_time_str.strip()}", "%Y-%m-%d %H.%M")
                else:
                    continue
            except:
                continue
                
            # Создаем название занятия
            title = f"{lesson_type} {discipline}"
            
            lessons.append({
                'title': title,
                'start_time': start_time,
                'end_time': end_time,
                'classroom': classroom,
                'group': group,
                'lesson_type': lesson_type,
                'discipline': discipline
            })
            
        return lessons
        
    except Exception as e:
        print(f"Error parsing Excel file: {e}")
        return []


@calendar_bp.route('/calendar')
@login_required
def calendar():
    groups = Group.query.filter_by(teacher_id=current_user.id).all()
    return render_template('calendar.html', groups=groups)


@calendar_bp.route('/api/schedule/events')
@login_required
def get_events():
    try:
        start = request.args.get('start')
        end = request.args.get('end')

        print(f"DEBUG: Getting events for teacher {current_user.id}")
        print(f"DEBUG: Start: {start}, End: {end}")

        # Получаем все события для текущего преподавателя
        query = Schedule.query.filter_by(teacher_id=current_user.id)
        
        # Применяем фильтры по дате, если они переданы
        if start:
            try:
                # Исправляем формат даты (заменяем пробел на +)
                start_fixed = start.replace(' ', '+')
                start_date = datetime.fromisoformat(start_fixed.replace('Z', '+00:00'))
                query = query.filter(Schedule.start_time >= start_date)
                print(f"DEBUG: Filtered by start date: {start_date}")
            except ValueError as e:
                print(f"DEBUG: Error parsing start date: {e}")
                
        if end:
            try:
                # Исправляем формат даты (заменяем пробел на +)
                end_fixed = end.replace(' ', '+')
                end_date = datetime.fromisoformat(end_fixed.replace('Z', '+00:00'))
                query = query.filter(Schedule.end_time <= end_date)
                print(f"DEBUG: Filtered by end date: {end_date}")
            except ValueError as e:
                print(f"DEBUG: Error parsing end date: {e}")

        all_events = query.all()
        print(f"DEBUG: Found {len(all_events)} events in database")

        events = []
        for event in all_events:
            # Для мероприятий (не занятий)
            if hasattr(event, 'is_event') and event.is_event:
                # Формируем время и место проведения
                start_time = event.start_time.strftime('%H:%M')
                end_time = event.end_time.strftime('%H:%M')
                time_info = f"{start_time}-{end_time}"
                if event.classroom:
                    time_info += f" • {event.classroom}"
                
                # Формируем заголовок для мероприятия
                event_type_names = {
                    'conference': 'Конференция',
                    'seminar': 'Семинар',
                    'meeting': 'Встреча',
                    'workshop': 'Мастер-класс',
                    'presentation': 'Презентация',
                    'other': 'Мероприятие'
                }
                event_type = getattr(event, 'event_type', 'other')
                event_type_name = event_type_names.get(event_type, 'Мероприятие')
                
                title_lines = [
                    f"📅 {event_type_name}",
                    time_info,
                    event.title
                ]
                description = getattr(event, 'description', '')
                if description:
                    title_lines.append(description)
                
                formatted_title = '\n'.join(title_lines)
                
                event_data = {
                    'id': event.id,
                    'title': formatted_title,
                    'start': event.start_time.isoformat(),
                    'end': event.end_time.isoformat(),
                    'color': event.color or '#dc3545',
                    'groupId': None,  # Мероприятия не привязаны к группам
                    'groupColor': event.color or '#dc3545',
                    'classroom': event.classroom,
                    'is_event': True,
                    'description': description,
                    'event_type': event_type
                }
                events.append(event_data)
                print(f"DEBUG: Event: {event_data}")
                continue
            
            # Для обычных занятий (существующая логика)
            group = Group.query.get(event.group_id)
            
            # Получаем аудиторию
            classroom = event.classroom or ""
            if not classroom:
                lesson = Lesson.query.filter_by(
                    group_id=event.group_id,
                    teacher_id=current_user.id,
                    topic=event.title
                ).first()
                if lesson and lesson.classroom:
                    classroom = lesson.classroom
            
            # Формируем время и аудиторию
            start_time = event.start_time.strftime('%H:%M')
            end_time = event.end_time.strftime('%H:%M')
            time_info = f"{start_time}-{end_time}"
            if classroom:
                time_info += f" • {classroom}"
            
            # Формируем новый формат отображения
            group_name = group.name if group else 'Неизвестная группа'
            discipline = event.title
            
            # Создаем многострочный заголовок
            title_lines = [
                group_name,
                time_info,
                discipline
            ]
            formatted_title = '\n'.join(title_lines)
            
            event_data = {
                'id': event.id,
                'title': formatted_title,
                'start': event.start_time.isoformat(),
                'end': event.end_time.isoformat(),
                'color': event.color or '#3788d8',
                'groupId': event.group_id,
                'groupColor': group.color if group else '#3788d8',
                'classroom': classroom,
                'is_event': False
            }
            events.append(event_data)
            print(f"DEBUG: Event: {event_data}")

        print(f"DEBUG: Returning {len(events)} events")
        return jsonify(events)
        
    except Exception as e:
        print(f"DEBUG: Error in get_events: {e}")
        return jsonify({'error': str(e)}), 500


@calendar_bp.route('/api/schedule/all-events')
@login_required
def get_all_events():
    """Получить все события преподавателя для отладки"""
    try:
        events = Schedule.query.filter_by(teacher_id=current_user.id).all()
        print(f"DEBUG: Found {len(events)} total events for teacher {current_user.id}")
        
        result = []
        for event in events:
            group = Group.query.get(event.group_id)
            result.append({
                'id': event.id,
                'title': event.title,
                'start': event.start_time.isoformat(),
                'end': event.end_time.isoformat(),
                'color': event.color or '#3788d8',
                'group_id': event.group_id,
                'group_name': group.name if group else 'Неизвестная группа',
                'teacher_id': event.teacher_id
            })
        
        return jsonify({
            'total_events': len(result),
            'events': result
        })
        
    except Exception as e:
        print(f"DEBUG: Error in get_all_events: {e}")
        return jsonify({'error': str(e)}), 500


@calendar_bp.route('/api/schedule/create-event', methods=['POST'])
@login_required
def create_event():
    """Создание мероприятия (не занятия)"""
    try:
        data = request.json
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        # Валидация обязательных полей
        required_fields = ['title', 'start', 'end']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Field {field} is required'}), 400

        # Создаем мероприятие в расписании
        event = Schedule(
            title=data['title'],
            start_time=datetime.fromisoformat(data['start']),
            end_time=datetime.fromisoformat(data['end']),
            group_id=None,  # Мероприятия не привязаны к группам
            color=data.get('color', '#dc3545'),
            classroom=data.get('location', ''),  # Используем поле classroom для места проведения
            teacher_id=current_user.id,
            is_event=True,  # Помечаем как мероприятие
            description=data.get('description', ''),
            event_type=data.get('event_type', 'other')
        )

        db.session.add(event)
        db.session.commit()

        return jsonify({
            'id': event.id, 
            'status': 'success',
            'message': 'Мероприятие успешно создано'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@calendar_bp.route('/api/schedule/create', methods=['POST'])
@login_required
def create_lesson():
    """Создание занятия (не мероприятия)"""
    try:
        data = request.json
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        # Валидация обязательных полей
        required_fields = ['title', 'start', 'end', 'group_id']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Field {field} is required'}), 400

        # Создаем событие в расписании
        event = Schedule(
            title=data['title'],
            start_time=datetime.fromisoformat(data['start']),
            end_time=datetime.fromisoformat(data['end']),
            group_id=data['group_id'],
            color=data.get('color', '#3788d8'),
            classroom=data.get('classroom', ''),
            teacher_id=current_user.id
        )

        db.session.add(event)
        db.session.flush()  # Получаем ID события

        # Создаем соответствующее занятие в журнале
        lesson = Lesson(
            date=event.start_time,
            group_id=event.group_id,
            topic=event.title,
            notes=f"Занятие создано из календаря. Время: {event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}",
            classroom=event.classroom,
            teacher_id=current_user.id
        )
        db.session.add(lesson)
        db.session.commit()

        return jsonify({
            'id': event.id, 
            'lesson_id': lesson.id,
            'status': 'success',
            'message': 'Занятие успешно создано'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@calendar_bp.route('/api/schedule/preview-excel', methods=['POST'])
@login_required
def preview_schedule_excel():
    """Предварительный просмотр Excel файла для настройки колонок"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Файл не найден'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Файл не выбран'}), 400
            
        if not file.filename.lower().endswith(('.xls', '.xlsx')):
            return jsonify({'error': 'Поддерживаются только Excel файлы (.xls, .xlsx)'}), 400
            
        # Сохраняем файл временно
        upload_folder = 'uploads'
        os.makedirs(upload_folder, exist_ok=True)
        
        # Определяем расширение файла
        file_extension = '.xlsx' if file.filename.lower().endswith('.xlsx') else '.xls'
        file_path = os.path.join(upload_folder, f"preview_{current_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_extension}")
        file.save(file_path)
        
        try:
            # Читаем Excel файл для предварительного просмотра
            if file_extension == '.xlsx':
                df = pd.read_excel(file_path, sheet_name=0, header=None, engine='openpyxl')
            else:
                df = pd.read_excel(file_path, sheet_name=0, header=None, engine='xlrd')
            
            # Получаем первые 10 строк для предварительного просмотра
            preview_data = []
            for index, row in df.iterrows():
                if index >= 10:  # Ограничиваем 10 строками
                    break
                preview_data.append([str(cell) if pd.notna(cell) else '' for cell in row])
            
            # Получаем информацию о колонках (первые 3 строки данных, пропуская заголовки)
            columns_info = []
            max_cols = min(len(df.columns), 20)  # Максимум 20 колонок
            for col_index in range(max_cols):
                col_data = []
                # Начинаем с первой строки данных (индекс 0), берем максимум 3 строки
                for row_index in range(min(3, len(df))):
                    cell_value = df.iloc[row_index, col_index] if col_index < len(df.columns) else None
                    col_data.append(str(cell_value) if pd.notna(cell_value) else '')
                columns_info.append(' | '.join(col_data[:3]))
            
            # Удаляем временный файл
            if os.path.exists(file_path):
                os.remove(file_path)
            
            return jsonify({
                'status': 'success',
                'preview': preview_data,
                'columns': columns_info
            })
            
        except Exception as e:
            print(f"DEBUG: Error reading Excel file: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({'error': f'Ошибка чтения файла: {str(e)}'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@calendar_bp.route('/api/schedule/upload-excel', methods=['POST'])
@login_required
def upload_schedule_excel():
    """Загрузка расписания из Excel файла с настройкой колонок"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Файл не найден'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Файл не выбран'}), 400
            
        if not file.filename.lower().endswith(('.xls', '.xlsx')):
            return jsonify({'error': 'Поддерживаются только Excel файлы (.xls, .xlsx)'}), 400
        
        # Получаем настройки колонок
        column_mapping = json.loads(request.form.get('column_mapping', '{}'))
        start_row = int(request.form.get('start_row', 6)) - 1  # Конвертируем в 0-based индекс
            
        # Сохраняем файл временно
        upload_folder = 'uploads'
        os.makedirs(upload_folder, exist_ok=True)
        
        # Определяем расширение файла
        file_extension = '.xlsx' if file.filename.lower().endswith('.xlsx') else '.xls'
        file_path = os.path.join(upload_folder, f"temp_schedule_{current_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_extension}")
        file.save(file_path)
        
        try:
            # Парсим Excel файл с настройками колонок
            lessons = parse_schedule_excel_with_mapping(file_path, column_mapping, start_row, file_extension)
            
            if not lessons:
                return jsonify({'error': 'Не удалось извлечь данные из файла'}), 400
                
            # Создаем группы, если их нет
            group_mapping = {}
            for lesson in lessons:
                group_names = lesson['group']
                # Обрабатываем группы, разделенные запятыми
                if ',' in group_names:
                    group_list = [g.strip() for g in group_names.split(',')]
                else:
                    group_list = [group_names.strip()]
                
                # Создаем или находим группы
                lesson_group_ids = []
                for group_name in group_list:
                    if group_name and group_name not in group_mapping:
                        # Ищем существующую группу
                        existing_group = Group.query.filter_by(
                            name=group_name, 
                            teacher_id=current_user.id
                        ).first()
                        
                        if existing_group:
                            group_mapping[group_name] = existing_group.id
                        else:
                            # Создаем новую группу
                            new_group = Group(
                                name=group_name,
                                course="Импортированная группа",
                                education_form="очная",
                                teacher_id=current_user.id,
                                color="#ffc107"  # Желтый цвет для импортированных групп
                            )
                            db.session.add(new_group)
                            db.session.flush()  # Получаем ID
                            group_mapping[group_name] = new_group.id
                    
                    if group_name in group_mapping:
                        lesson_group_ids.append(group_mapping[group_name])
                
                # Обновляем lesson с правильными ID групп
                lesson['group_ids'] = lesson_group_ids
                lesson['group_names'] = group_list
            
            # Создаем занятия в расписании и журнале
            created_lessons = []
            for lesson in lessons:
                # Создаем занятие для каждой группы
                for group_id in lesson['group_ids']:
                    # Проверяем, не существует ли уже такое занятие для этой группы
                    existing_schedule = Schedule.query.filter_by(
                        title=lesson['title'],
                        start_time=lesson['start_time'],
                        end_time=lesson['end_time'],
                        group_id=group_id,
                        teacher_id=current_user.id
                    ).first()
                    
                    if existing_schedule:
                        continue
                        
                    schedule = Schedule(
                        title=lesson['title'],
                        start_time=lesson['start_time'],
                        end_time=lesson['end_time'],
                        group_id=group_id,
                        classroom=lesson['classroom'],
                        color="#ffc107",  # Желтый цвет для импортированных занятий
                        teacher_id=current_user.id
                    )
                    
                    db.session.add(schedule)
                    db.session.flush()  # Получаем ID события
                    
                    # Создаем соответствующее занятие в журнале
                    journal_lesson = Lesson(
                        date=lesson['start_time'],
                        group_id=group_id,
                        topic=lesson['title'],
                        notes=f"Занятие импортировано из Excel. Время: {lesson['start_time'].strftime('%H:%M')} - {lesson['end_time'].strftime('%H:%M')}",
                        classroom=lesson['classroom'],
                        teacher_id=current_user.id
                    )
                    db.session.add(journal_lesson)
                    
                    created_lessons.append(f"{lesson['title']} - {', '.join(lesson['group_names'])}")
            
            db.session.commit()
            
            # Удаляем временный файл
            try:
                os.remove(file_path)
            except:
                pass
                
            return jsonify({
                'status': 'success',
                'message': f'Успешно загружено {len(created_lessons)} занятий',
                'lessons_count': len(created_lessons),
                'lessons': created_lessons[:10]  # Показываем первые 10 занятий
            })
            
        except Exception as e:
            db.session.rollback()
            # Удаляем временный файл в случае ошибки
            try:
                os.remove(file_path)
            except:
                pass
            return jsonify({'error': f'Ошибка при обработке файла: {str(e)}'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Ошибка загрузки: {str(e)}'}), 500


@calendar_bp.route('/api/schedule/update/<int:event_id>', methods=['PUT'])
@login_required
def update_event(event_id):
    try:
        event = Schedule.query.get_or_404(event_id)

        if event.teacher_id != current_user.id:
            return jsonify({'error': 'Unauthorized'}), 403

        data = request.json
        
        # Сохраняем старые значения для поиска занятия
        old_title = event.title
        old_start_time = event.start_time
        old_group_id = event.group_id
        
        # Обновляем событие в расписании
        if 'title' in data:
            event.title = data['title']
        if 'start' in data:
            event.start_time = datetime.fromisoformat(data['start'])
        if 'end' in data:
            event.end_time = datetime.fromisoformat(data['end'])
        if 'color' in data:
            event.color = data['color']
        if 'classroom' in data:
            event.classroom = data['classroom']
        if 'group_id' in data:
            event.group_id = data['group_id']

        # Находим и обновляем соответствующее занятие в журнале
        # Ищем по старому названию, дате и группе, чтобы найти правильное занятие
        lesson = Lesson.query.filter_by(
            group_id=old_group_id,
            teacher_id=current_user.id,
            topic=old_title,
            date=old_start_time
        ).first()
        
        if lesson:
            lesson.date = event.start_time
            lesson.topic = event.title
            lesson.group_id = event.group_id  # Обновляем группу занятия
            lesson.notes = f"Занятие обновлено из календаря. Время: {event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}"
            lesson.classroom = event.classroom
        else:
            # Если занятие не найдено, создаем новое
            lesson = Lesson(
                date=event.start_time,
                group_id=event.group_id,
                topic=event.title,
                notes=f"Занятие создано из календаря. Время: {event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}",
                classroom=event.classroom,
                teacher_id=current_user.id
            )
            db.session.add(lesson)

        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Занятие успешно обновлено'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@calendar_bp.route('/api/schedule/delete/<int:event_id>', methods=['DELETE'])
def delete_event(event_id):
    try:
        print(f"DEBUG: Attempting to delete event {event_id}")
        print(f"DEBUG: Current user: {current_user}")
        print(f"DEBUG: Current user ID: {current_user.id if current_user.is_authenticated else 'Not authenticated'}")
        
        if not current_user.is_authenticated:
            print("DEBUG: User not authenticated")
            return jsonify({'error': 'Authentication required'}), 401
        
        event = Schedule.query.get_or_404(event_id)
        print(f"DEBUG: Found event: {event.title} for group {event.group_id}")

        if event.teacher_id != current_user.id:
            print(f"DEBUG: Unauthorized - event belongs to teacher {event.teacher_id}, current user is {current_user.id}")
            return jsonify({'error': 'Unauthorized'}), 403

        # Находим и удаляем соответствующее занятие в журнале
        # Ищем по точному совпадению названия, даты и группы
        lesson = Lesson.query.filter_by(
            group_id=event.group_id,
            teacher_id=current_user.id,
            topic=event.title,
            date=event.start_time
        ).first()
        
        if lesson:
            print(f"DEBUG: Found corresponding lesson {lesson.id}, deleting attendance records")
            # Удаляем все записи посещаемости для этого занятия
            try:
                attendance_count = Attendance.query.filter_by(lesson_id=lesson.id).count()
                print(f"DEBUG: Found {attendance_count} attendance records to delete")
                if attendance_count > 0:
                    Attendance.query.filter_by(lesson_id=lesson.id).delete()
                db.session.delete(lesson)
                print(f"DEBUG: Deleted lesson and {attendance_count} attendance records")
            except Exception as e:
                print(f"DEBUG: Error deleting lesson/attendance: {str(e)}")
                # Продолжаем удаление события даже если есть проблемы с журналом
        else:
            print("DEBUG: No corresponding lesson found in journal")

        # Удаляем событие из расписания
        try:
            db.session.delete(event)
            db.session.commit()
            print(f"DEBUG: Successfully deleted event {event_id}")
            return jsonify({'status': 'success', 'message': 'Занятие успешно удалено'})
        except Exception as e:
            print(f"DEBUG: Error deleting event from database: {str(e)}")
            db.session.rollback()
            raise e
        
    except Exception as e:
        print(f"DEBUG: Error deleting event {event_id}: {str(e)}")
        import traceback
        print(f"DEBUG: Full traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'error': f'Ошибка при удалении занятия: {str(e)}'}), 500


@calendar_bp.route('/api/schedule/suggest-slot', methods=['POST'])
@login_required
def suggest_slot():
    data = request.json
    duration = data.get('duration', 90)
    preferred_days = data.get('preferred_days', ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])

    existing_events = Schedule.query.filter_by(teacher_id=current_user.id).all()

    schedule_data = [{
        'day': event.start_time.strftime('%A'),
        'start': event.start_time.strftime('%H:%M'),
        'end': event.end_time.strftime('%H:%M')
    } for event in existing_events]

    suggestion = ai.suggest_schedule_slot(schedule_data, duration)

    return jsonify(suggestion)


@calendar_bp.route('/api/schedule/optimize', methods=['POST'])
@login_required
def optimize_schedule():
    events = Schedule.query.filter_by(teacher_id=current_user.id).all()

    time_slots = []
    for hour in range(8, 20):
        for day in range(5):
            slot_start = datetime.now().replace(hour=hour, minute=0, second=0)
            slot_start += timedelta(days=day - datetime.now().weekday())

            is_free = True
            for event in events:
                if (event.start_time <= slot_start < event.end_time):
                    is_free = False
                    break

            if is_free:
                time_slots.append({
                    'day': slot_start.strftime('%A'),
                    'time': slot_start.strftime('%H:%M'),
                    'date': slot_start.isoformat()
                })

    return jsonify({'free_slots': time_slots[:10]})


@calendar_bp.route('/api/schedule/conflicts')
@login_required
def check_conflicts():
    events = Schedule.query.filter_by(teacher_id=current_user.id).order_by(Schedule.start_time).all()

    conflicts = []
    for i in range(len(events) - 1):
        if events[i].end_time > events[i + 1].start_time:
            conflicts.append({
                'event1': {'id': events[i].id, 'title': events[i].title},
                'event2': {'id': events[i + 1].id, 'title': events[i + 1].title},
                'overlap_minutes': (events[i].end_time - events[i + 1].start_time).seconds // 60
            })

    return jsonify({'conflicts': conflicts})


@calendar_bp.route('/api/schedule/sync-to-journal', methods=['POST'])
@login_required
def sync_schedule_to_journal():
    """Синхронизирует все занятия из календаря в журнал"""
    try:
        # Получаем все события из календаря
        events = Schedule.query.filter_by(teacher_id=current_user.id).all()
        
        synced_count = 0
        created_count = 0
        
        for event in events:
            # Проверяем, есть ли уже такое занятие в журнале
            existing_lesson = Lesson.query.filter_by(
                group_id=event.group_id,
                teacher_id=current_user.id,
                topic=event.title,
                date=event.start_time
            ).first()
            
            if not existing_lesson:
                # Создаем новое занятие в журнале
                lesson = Lesson(
                    date=event.start_time,
                    group_id=event.group_id,
                    topic=event.title,
                    notes=f"Занятие синхронизировано из календаря. Время: {event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}",
                    classroom=event.classroom,
                    teacher_id=current_user.id
                )
                db.session.add(lesson)
                created_count += 1
            else:
                # Обновляем существующее занятие
                existing_lesson.classroom = event.classroom
                existing_lesson.notes = f"Занятие синхронизировано из календаря. Время: {event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}"
                synced_count += 1
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Синхронизация завершена. Создано: {created_count}, обновлено: {synced_count}',
            'created': created_count,
            'updated': synced_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@calendar_bp.route('/api/schedule/sync-status')
@login_required
def get_sync_status():
    """Возвращает статус синхронизации между календарем и журналом"""
    try:
        # Получаем все события из календаря
        calendar_events = Schedule.query.filter_by(teacher_id=current_user.id).all()
        
        # Получаем все занятия из журнала
        journal_lessons = Lesson.query.filter_by(teacher_id=current_user.id).all()
        
        # Подсчитываем недостающие занятия
        missing_count = 0
        for event in calendar_events:
            existing_lesson = Lesson.query.filter_by(
                group_id=event.group_id,
                teacher_id=current_user.id,
                topic=event.title,
                date=event.start_time
            ).first()
            if not existing_lesson:
                missing_count += 1
        
        return jsonify({
            'calendar_events': len(calendar_events),
            'journal_lessons': len(journal_lessons),
            'missing_lessons': missing_count,
            'sync_required': missing_count > 0
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@calendar_bp.route('/api/schedule/export-excel')
@login_required
def export_schedule_excel():
    """Экспорт расписания в Excel файл"""
    try:
        # Получаем все события для текущего преподавателя
        events = Schedule.query.filter_by(teacher_id=current_user.id).order_by(Schedule.start_time).all()
        
        if not events:
            return jsonify({'error': 'Нет занятий для экспорта'}), 400
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание занятий"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Заголовки
        headers = [
            'Дата', 'День недели', 'Время начала', 'Время окончания', 
            'Группа', 'Дисциплина', 'Аудитория', 'Преподаватель'
        ]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Записываем данные
        for row, event in enumerate(events, 2):
            group = Group.query.get(event.group_id)
            group_name = group.name if group else 'Неизвестная группа'
            
            # Форматируем дату и время
            date_str = event.start_time.strftime('%d.%m.%Y')
            weekday = event.start_time.strftime('%A')
            # Переводим день недели на русский
            weekdays_ru = {
                'Monday': 'Понедельник',
                'Tuesday': 'Вторник', 
                'Wednesday': 'Среда',
                'Thursday': 'Четверг',
                'Friday': 'Пятница',
                'Saturday': 'Суббота',
                'Sunday': 'Воскресенье'
            }
            weekday_ru = weekdays_ru.get(weekday, weekday)
            
            start_time = event.start_time.strftime('%H:%M')
            end_time = event.end_time.strftime('%H:%M')
            
            # Записываем данные в строку
            row_data = [
                date_str,
                weekday_ru,
                start_time,
                end_time,
                group_name,
                event.title,
                event.classroom or '',
                current_user.username
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment
        
        # Автоподбор ширины колонок
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(events) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Замораживаем первую строку
        ws.freeze_panes = 'A2'
        
        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Генерируем имя файла
        current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"schedule_{current_user.username}_{current_date}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Ошибка при экспорте: {str(e)}'}), 500


@calendar_bp.route('/api/schedule/create-recurring', methods=['POST'])
@login_required
def create_recurring_lessons():
    """Создание повторяющихся занятий"""
    try:
        data = request.json
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        # Валидация обязательных полей
        required_fields = ['title', 'group_id', 'start_time', 'end_time', 'start_date', 'end_date', 'days_of_week']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Field {field} is required'}), 400

        # Парсим даты
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
        
        # Парсим время
        start_time_str = data['start_time']
        end_time_str = data['end_time']
        
        # Получаем дни недели (1=Понедельник, 2=Вторник, ..., 6=Суббота)
        days_of_week = data['days_of_week']
        
        if not isinstance(days_of_week, list) or len(days_of_week) == 0:
            return jsonify({'error': 'At least one day of week must be selected'}), 400
        
        # Проверяем, что дата начала не позже даты окончания
        if start_date > end_date:
            return jsonify({'error': 'Start date cannot be later than end date'}), 400
        
        # Создаем занятия для каждого выбранного дня недели в указанном периоде
        created_lessons = []
        current_date = start_date
        
        while current_date <= end_date:
            # Проверяем, является ли текущая дата одним из выбранных дней недели
            # Python weekday(): Monday=0, Tuesday=1, ..., Sunday=6
            # Наши дни: Monday=1, Tuesday=2, ..., Saturday=6
            weekday = current_date.weekday() + 1  # Конвертируем в наш формат
            
            if weekday in days_of_week:
                # Создаем дату и время для занятия
                lesson_datetime_start = datetime.combine(current_date, datetime.strptime(start_time_str, '%H:%M').time())
                lesson_datetime_end = datetime.combine(current_date, datetime.strptime(end_time_str, '%H:%M').time())
                
                # Проверяем, не существует ли уже такое занятие
                existing_schedule = Schedule.query.filter_by(
                    title=data['title'],
                    start_time=lesson_datetime_start,
                    end_time=lesson_datetime_end,
                    group_id=data['group_id'],
                    teacher_id=current_user.id
                ).first()
                
                if not existing_schedule:
                    # Создаем событие в расписании
                    schedule = Schedule(
                        title=data['title'],
                        start_time=lesson_datetime_start,
                        end_time=lesson_datetime_end,
                        group_id=data['group_id'],
                        classroom=data.get('classroom', ''),
                        color=data.get('color', '#3788d8'),
                        teacher_id=current_user.id
                    )
                    
                    db.session.add(schedule)
                    db.session.flush()  # Получаем ID события
                    
                    # Создаем соответствующее занятие в журнале
                    lesson = Lesson(
                        date=lesson_datetime_start,
                        group_id=data['group_id'],
                        topic=data['title'],
                        notes=f"Повторяющееся занятие. Время: {start_time_str} - {end_time_str}",
                        classroom=data.get('classroom', ''),
                        teacher_id=current_user.id
                    )
                    db.session.add(lesson)
                    
                    created_lessons.append({
                        'date': current_date.strftime('%Y-%m-%d'),
                        'day_name': current_date.strftime('%A'),
                        'time': f"{start_time_str} - {end_time_str}"
                    })
            
            # Переходим к следующему дню
            current_date += timedelta(days=1)
        
        # Сохраняем все изменения
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully created {len(created_lessons)} recurring lessons',
            'lessons_created': len(created_lessons),
            'lessons': created_lessons[:10]  # Показываем первые 10 занятий
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500